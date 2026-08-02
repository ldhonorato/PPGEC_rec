import csv
import io
import unicodedata
from pathlib import Path

from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction
from django.db.models import Q

from .models import Aluno, TrajetoriaAcademica, User, validar_cpf_brasileiro


COLUNAS_OBRIGATORIAS = {"nome", "cpf", "email", "orientador"}


def _normalizar_coluna(valor):
    texto = unicodedata.normalize("NFKD", str(valor or ""))
    texto = "".join(char for char in texto if not unicodedata.combining(char))
    return texto.strip().lower().replace("-", "").replace("_", "").replace(" ", "")


def _linhas_csv(arquivo):
    conteudo = arquivo.read()
    for encoding in ("utf-8-sig", "latin-1"):
        try:
            texto = conteudo.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    amostra = texto[:4096]
    try:
        dialect = csv.Sniffer().sniff(amostra, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    return list(csv.reader(io.StringIO(texto), dialect))


def _linhas_excel(arquivo, extensao):
    if extensao == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(arquivo, read_only=True, data_only=True)
        try:
            return [list(linha) for linha in workbook.active.iter_rows(values_only=True)]
        finally:
            workbook.close()

    try:
        import xlrd
    except ImportError as exc:
        raise ValidationError("A leitura de arquivos XLS não está disponível no servidor.") from exc
    workbook = xlrd.open_workbook(file_contents=arquivo.read())
    planilha = workbook.sheet_by_index(0)
    return [planilha.row_values(indice) for indice in range(planilha.nrows)]


def ler_planilha_ingressantes(arquivo):
    extensao = Path(arquivo.name).suffix.lower()
    linhas = _linhas_csv(arquivo) if extensao == ".csv" else _linhas_excel(arquivo, extensao)
    if not linhas:
        raise ValidationError("A planilha está vazia.")

    aliases = {"nome": "nome", "cpf": "cpf", "email": "email", "orientador": "orientador"}
    cabecalho = [aliases.get(_normalizar_coluna(valor)) for valor in linhas[0]]
    encontradas = {coluna for coluna in cabecalho if coluna}
    faltantes = COLUNAS_OBRIGATORIAS - encontradas
    if faltantes:
        raise ValidationError("Colunas obrigatórias ausentes: " + ", ".join(sorted(faltantes)) + ".")

    registros = []
    for numero, valores in enumerate(linhas[1:], start=2):
        registro = {
            coluna: str(valores[indice] or "").strip()
            for indice, coluna in enumerate(cabecalho)
            if coluna and indice < len(valores)
        }
        if any(registro.values()):
            registros.append((numero, registro))
    return registros


def importar_ingressantes(*, arquivo, nivel_curso, ingresso):
    resultados = []
    for numero, dados in ler_planilha_ingressantes(arquivo):
        nome = dados.get("nome", "").strip()
        cpf = "".join(char for char in dados.get("cpf", "") if char.isdigit())
        email = dados.get("email", "").strip().lower()
        orientador_nome = dados.get("orientador", "").strip()
        resultado = {"linha": numero, "nome": nome or "(sem nome)", "cadastrado": False}

        faltantes = [campo for campo in ("nome", "cpf", "email") if not {"nome": nome, "cpf": cpf, "email": email}[campo]]
        if faltantes:
            resultado["motivo"] = "Campos não preenchidos: " + ", ".join(faltantes) + "."
            resultados.append(resultado)
            continue
        if not validar_cpf_brasileiro(cpf):
            resultado["motivo"] = "CPF inválido."
            resultados.append(resultado)
            continue

        existente = Aluno.objects.filter(
            Q(cpf=cpf) | Q(nome__iexact=nome, trajetorias__ingresso=ingresso)
        ).distinct().first()
        if existente:
            resultado["motivo"] = f"O aluno já possui cadastro ({existente.email})."
            resultados.append(resultado)
            continue
        if User.objects.filter(email__iexact=email).exists():
            resultado["motivo"] = "O e-mail já possui cadastro."
            resultados.append(resultado)
            continue

        orientador = None
        if nivel_curso != Aluno.NivelCurso.ALUNO_ESPECIAL:
            orientadores = User.objects.filter(
                tipo_usuario=User.TipoUsuario.DOCENTE,
                nome__iexact=orientador_nome,
                is_active=True,
            )
            if not orientador_nome or orientadores.count() != 1:
                resultado["motivo"] = "Orientador não encontrado de forma única."
                resultados.append(resultado)
                continue
            orientador = orientadores.first()

        try:
            with transaction.atomic():
                aluno = Aluno.objects.create_user(
                    email=email,
                    password=None,
                    nome=nome,
                    cpf=cpf,
                    status_aluno=Aluno.StatusAluno.ATIVO,
                )
                TrajetoriaAcademica.objects.create(
                    aluno=aluno,
                    nivel_curso=nivel_curso,
                    status=TrajetoriaAcademica.Status.ATIVA,
                    ingresso=ingresso,
                    orientador=orientador,
                )
        except (ValidationError, IntegrityError) as exc:
            mensagens = getattr(exc, "messages", None)
            resultado["motivo"] = "; ".join(mensagens) if mensagens else "Dados inválidos ou duplicados."
        else:
            resultado["cadastrado"] = True
            resultado["motivo"] = "Cadastrado com sucesso."
        resultados.append(resultado)
    return resultados
